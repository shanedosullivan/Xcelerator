from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from xlrd import open_workbook
from openpyxl.styles import Font, Color, Alignment, colors

class Converter:

    def convert_and_save_excel(self, destination, fileName):
        valuesToWrite = []
        descriptionsToWrite = []
        cells = []

        wb = open_workbook(fileName)
        for sheet in wb.sheets():
            noRows = sheet.nrows
            for row in range(1, noRows):
                rowValue = sheet.cell(row, 2).value
                if rowValue is not None and rowValue != "":
                    valuesToWrite.append(str(rowValue))
                    descriptionsToWrite.append(sheet.cell(row, 0).value)

        newWb = Workbook()
        ws1 = newWb.active
        ws1.title = "Sheet1"
        ws1.merge_cells('A1:B1')
        alignment = Alignment(horizontal='center')

        ws1['A1'] = "Rand York Costs"
        ws1['B2'] = "Value"
        ws1['A2'] = "Description"

        titleCell = ws1['A1']
        valueTitlecell = ws1['B2']
        descTitleCell = ws1['A2']
        cells.append(titleCell)
        cells.append(valueTitlecell)
        cells.append(descTitleCell)

        for myCell in cells:
            font = Font(color=colors.RED, bold=True)
            myCell.font = font
            myCell.alignment = alignment

        rowCounter = 1;

        for row in range(3, len(valuesToWrite)):
            val = ws1.cell(column=2, row=row, value=valuesToWrite[rowCounter].format(get_column_letter(2)))
            desc = ws1.cell(column=1, row=row, value=descriptionsToWrite[rowCounter].format(get_column_letter(1)))
            rowCounter += 1

        for col in ws1.columns:
             maxLength = 0
             column = col[0].column # Get the column name
             for cell in col:
                 try: # Necessary to avoid error on empty cells
                     if len(str(cell.value)) > maxLength:
                         maxLength = len(cell.value)
                 except:
                     pass
             adjustedWidth = (maxLength + 2) * 1.2
             ws1.column_dimensions[column].width = adjustedWidth

        newWb.save(filename = destination)
